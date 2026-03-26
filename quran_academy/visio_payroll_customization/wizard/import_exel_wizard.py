from odoo import models, fields, api
import base64
import io
from openpyxl import load_workbook
from datetime import datetime, timedelta
from odoo.exceptions import ValidationError


class ExcelImportWizard(models.TransientModel):
    _name = 'excel.import.wizard'
    _description = 'Excel Import Wizard'

    file = fields.Binary(string="Excel File")
    download_sample_url = fields.Char('Download Sample URL', compute="_compute_download_sample_url", store=True)

    @api.depends('month')
    def _compute_download_sample_url(self):
        for rec in self:
            if rec.month:
                rec.download_sample_url = f"/download/excel/file/{rec.month}"

    def action_download_sample_url(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_url',
            'url': self.download_sample_url or '#',  # Your URL here
            'target': 'new',  # 'new' => open in new tab/window
        }

    def _default_month(self):
        return datetime.today().replace(day=1)

    month = fields.Date(string="Month", required=True , default=_default_month)
    record_ids = fields.Many2many('hr.payslip', string="Selected Payslips", readonly=True)


    def import_excel_file(self):
        self.ensure_one()
        if not self.file:
            raise ValidationError("Please Upload Excel File!")
        try:
            file_data = base64.b64decode(self.file)
            excel_file = io.BytesIO(file_data)
            workbook = load_workbook(excel_file, data_only=True)
            sheet = workbook.active

            required_columns = [
                # Basic Salary Tab
                'Worker Code',
                'Employee Name',
                'Designation',
                'Rate Per Month',
                'Other Duty',
                'Comments',

                # Allowances Tab
                'Medical',
                'Ext-Arrears-Night',
                'Daily Wages',
                # 'Rent Allowance',
                'Other Deduction',
                'Special Allowance',

                'Leave in cash',
                'Over Time',
                'Without Pay',
                'Attendance Star',
                'One Time Deduction',

                # Deductions Tab
                'Loan Recovery',
                'Advance Salary',
                'Income Tax',
                'Provident Fund',
                'Mess Bill',
                'House Rent',
                'Vehicle Charges',
                'Electricity/Gas Bill',

                'Arrear EOBI Employer',
                'Arrear EOBI Worker',
                'EOBI Worker',
                'EOBI Employer',

                'EOBI',
                'Leave Encash 2/3',
                'Paid By',
            ]
            headers = [cell.value for cell in sheet[1]]

            for column in required_columns:
                if column not in headers:
                    raise ValidationError(f"Missing required column: {column}")

            header_index_map = {header: index for index, header in enumerate(headers)}
            excel_mapping = {}

            for row in sheet.iter_rows(min_row=2, values_only=True):
                worker_code = str(row[header_index_map['Worker Code']])
                print(worker_code)
                if worker_code:
                    excel_mapping[worker_code] = {

                        'employee_job_title': row[header_index_map['Designation']],
                        'rate_per_month': row[header_index_map['Rate Per Month']],
                        'other_duty': row[header_index_map['Other Duty']],
                        'comments': row[header_index_map['Comments']],

                        'medical': row[header_index_map['Medical']],
                        'ext_arrear_night': row[header_index_map['Ext-Arrears-Night']],
                        'daily_wages': row[header_index_map['Daily Wages']],
                        'special_allowance_qa': row[header_index_map['Special Allowance']],

                        # 'rent_allow': row[header_index_map['Rent Allowance']],
                        'other_deduction': row[header_index_map['Other Deduction']],

                        'leave_in_cash': row[header_index_map['Leave in cash']],
                        'over_time': row[header_index_map['Over Time']],
                        'wo_pay': row[header_index_map['Without Pay']],
                        'attendance_star': row[header_index_map['Attendance Star']],
                        'one_time_deduction': row[header_index_map['One Time Deduction']],

                        'loan_recovery': row[header_index_map['Loan Recovery']],
                        'adv_salary': row[header_index_map['Advance Salary']],

                        'income_tax': row[header_index_map['Income Tax']],
                        'prov_fund': row[header_index_map['Provident Fund']],

                        'tel_bill': row[header_index_map['Mess Bill']],
                        'h_rent': row[header_index_map['House Rent']],
                        'veh_charges': row[header_index_map['Vehicle Charges']],
                        'elec_gas_bill': row[header_index_map['Electricity/Gas Bill']],

                        'eobi_worker': row[header_index_map['EOBI Worker']],
                        'emp_eobi': row[header_index_map['EOBI Employer']],
                        'eobi_emp_diff': row[header_index_map['Arrear EOBI Employer']],
                        'eobi_worker_diff': row[header_index_map['Arrear EOBI Worker']],

                        'eobi': row[header_index_map['EOBI']],
                        'leave_encash': row[header_index_map['Leave Encash 2/3']],
                        'paid_by': row[header_index_map['Paid By']],
                    }


            selected_month_start = self.month.replace(day=1)
            next_month = self.month.replace(day=28) + timedelta(days=4)
            selected_month_end = next_month.replace(day=1) - timedelta(days=1)

            payslips = self.env['hr.payslip'].search([
                ('date_from', '>=', selected_month_start),
                ('date_to', '<=', selected_month_end),
                ('state', '=', 'draft')
            ])

            for record in payslips:
                if record.worker_code in excel_mapping:
                    record_data = excel_mapping[record.worker_code]
                    update_data = {}

                    if record_data['rate_per_month'] is not None:
                        update_data['rate_per_month'] = record_data['rate_per_month']
                    if record_data['other_duty'] is not None:
                        update_data['other_duty'] = record_data['other_duty']
                    if record_data['comments'] is not None:
                        update_data['comments'] = record_data['comments']

                    if record_data['medical'] is not None:
                        update_data['medical'] = record_data['medical']
                    if record_data['ext_arrear_night'] is not None:
                        update_data['ext_arrear_night'] = record_data['ext_arrear_night']
                    if record_data['daily_wages'] is not None:
                        update_data['daily_wages'] = record_data['daily_wages']
                    if record_data['special_allowance_qa'] is not None:
                        update_data['special_allowance_qa'] = record_data['special_allowance_qa']
                    # if record_data['rent_allow'] is not None:
                    #     update_data['rent_allow'] = record_data['rent_allow']
                    if record_data['other_deduction'] is not None:
                        update_data['other_deduction'] = record_data['other_deduction']

                    if record_data['leave_in_cash'] is not None:
                        update_data['leave_in_cash'] = record_data['leave_in_cash']
                    if record_data['over_time'] is not None:
                        update_data['over_time'] = record_data['over_time']
                    if record_data['wo_pay'] is not None:
                        update_data['wo_pay'] = record_data['wo_pay']
                    if record_data['attendance_star'] is not None:
                        update_data['attendance_star'] = record_data['attendance_star']
                    if record_data['one_time_deduction'] is not None:
                        update_data['one_time_deduction'] = record_data['one_time_deduction']

                    if record_data['loan_recovery'] is not None:
                        update_data['loan_recovery'] = record_data['loan_recovery']
                    if record_data['adv_salary'] is not None:
                        update_data['adv_salary'] = record_data['adv_salary']

                    if record_data['income_tax'] is not None:
                        update_data['income_tax'] = record_data['income_tax']
                    if record_data['prov_fund'] is not None:
                        update_data['prov_fund'] = record_data['prov_fund']

                    if record_data['tel_bill'] is not None:
                        if record_data['tel_bill']:
                            update_data['tel_bill'] = record_data['tel_bill']
                        else:
                            update_data['tel_bill'] = record_data['special_allowance_qa']
                    if record_data['h_rent'] is not None:
                        update_data['h_rent'] = record_data['h_rent']
                    if record_data['veh_charges'] is not None:
                        update_data['veh_charges'] = record_data['veh_charges']
                    if record_data['elec_gas_bill'] is not None:
                        update_data['elec_gas_bill'] = record_data['elec_gas_bill']

                    if record_data['eobi_worker'] is not None:
                        update_data['eobi_worker'] = record_data['eobi_worker']

                    if record_data['emp_eobi'] is not None:
                        update_data['emp_eobi'] = record_data['emp_eobi']

                    if record_data['eobi_emp_diff'] is not None:
                        update_data['eobi_emp_diff'] = record_data['eobi_emp_diff']

                    if record_data['eobi_worker_diff'] is not None:
                        update_data['eobi_worker_diff'] = record_data['eobi_worker_diff']

                    if record_data['eobi'] is not None:
                        update_data['eobi'] = record_data['eobi']
                    if record_data['leave_encash'] is not None:
                        update_data['leave_encash'] = record_data['leave_encash']
                    if record_data['paid_by'] is not None:
                        update_data['paid_by'] = record_data['paid_by']

                    if update_data:
                        print("updating data")
                        record.write(update_data)

            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': 'Success!',
                    'message': 'Payslips have been successfully updated.',
                    'next': {'type': 'ir.actions.client', 'tag': 'reload'},
                    'sticky': False,
                    'type': 'success',
                }
            }

        except Exception as e:
            raise ValidationError(f"Failed to process the Excel file: {str(e)}")
