import openpyxl
from numpy.ma.core import append
from odoo import models, fields, api
import logging

_logger = logging.getLogger(__name__)


class ResCountryUpdate(models.Model):
    _inherit = "res.country"

    # lims_code = fields.Char(string="LIMS Code")

    @api.model
    def update_countries_from_excel(self):
        file_path = "custom/addons/visio_tti_so_customize/update_data/country.xlsx"
        file_path = "src/user/visio_tti_so_customize/update_data/country.xlsx"

        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            headers = [str(cell.value or "").strip() for cell in sheet[1]]
            if "Name" not in headers or "Lims Code" not in headers:
                print("Error: Must have 'Name' and 'Lims Code' columns.")
                return

            name_idx = headers.index("Name")
            code_idx = headers.index("Lims Code")

            for row in sheet.iter_rows(min_row=2, values_only=True):
                country_name = (str(row[name_idx]).strip()
                                if row[name_idx] else "")
                lims_code = (str(row[code_idx]).strip()
                             if row[code_idx] else "")

                if not country_name:
                    print("Skipping empty country name.")
                    continue

                existing = self.env["res.country"].search(
                    [("name", "=", country_name)], limit=1
                )
                if existing:
                    existing.write({"lims_code": lims_code})
                    print(f"Updated {country_name} with LIMS Code: {lims_code}")
                    continue

                unique_code = "XX"
                for start in range(len(country_name) - 1):
                    code_candidate = country_name[start:start + 2].upper()
                    if not code_candidate.isalpha():
                        continue
                    existing_code = self.env["res.country"].search(
                        [("code", "=", code_candidate)], limit=1
                    )
                    if not existing_code:
                        unique_code = code_candidate
                        break

                new_country = self.env["res.country"].create({
                    "name": country_name,
                    "lims_code": lims_code,
                    "code": unique_code
                })
                _logger.info(f"Created new country {country_name} with Code: {unique_code} and LIMS Code: {lims_code}")
            _logger.info("Country update process completed.")
        except Exception as e:
            _logger.error(f"Error processing Excel file: {e}")



class ResPartnerUpdate(models.Model):
    _inherit = "res.partner"

    @api.model
    def update_brands_with_buyers(self):
        file_path = "custom/addons/visio_tti_so_customize/update_data/brand.xlsx"
        file_path = "src/user/visio_tti_so_customize/update_data/brand.xlsx"

        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]

            if "brandcode" not in headers or "brandname" not in headers or "buyercode" not in headers:
                print("Error: The sheet must have 'brandcode', 'brandname', and 'buyercode' columns.")
                return

            brandcode_idx = headers.index("brandcode")
            brandname_idx = headers.index("brandname")
            buyercode_idx = headers.index("buyercode")

            for row in sheet.iter_rows(min_row=2, values_only=True):
                brand_code = str(row[brandcode_idx]).strip() if row[brandcode_idx] else ""
                brand_name = str(row[brandname_idx]).strip() if row[brandname_idx] else ""
                buyer_code = str(row[buyercode_idx]).strip() if row[buyercode_idx] else ""

                if not brand_code or not brand_name or not buyer_code:
                    print(f"Skipping row with missing values: {row}")
                    continue

                brand = self.env["res.partner"].search([
                    ("code", "=", brand_code),
                    ("tti_company_category", "=", "brand")
                ], limit=1)

                if not brand:
                    print(f"Brand not found: {brand_name} ({brand_code}), skipping.")
                    continue

                buyer = self.env["res.partner"].search([
                    ("code", "=", buyer_code),
                    ("tti_company_category", "=", "buyer")
                ], limit=1)

                if not buyer:
                    print(f"Buyer not found for {buyer_code}, skipping assignment.")
                    continue

                brand.write({"parent_id": buyer.id})
                print(f"Assigned Buyer ({buyer_code}) to Brand ({brand_name} - {brand_code})")
            print("Brand-Buyer update process completed.")
        except Exception as e:
            _logger.error(f"Error processing Excel file: {e}")
            # print(f"Error processing Excel file: {e}")

    @api.model
    def update_buyers_with_packages(self):
        # file_path = "custom_addons/2025/03_Mar/17_03_2025/1_tti_testing_v18/tti_testing/visio_tti_so_customize/update_data/buyer_wise_packages.xlsx"
        file_path = "src/user/visio_tti_so_customize/update_data/buyer_wise_packages.xlsx"

        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]

            if "buyercode" not in headers or "buyername" not in headers or "packagecode" not in headers:
                print("Error: The sheet must have required columns columns.")
                return

            buyercode_idx = headers.index("buyercode")
            buyername_idx = headers.index("buyername")
            packagename_idx = headers.index("packagename")
            packagecode_idx = headers.index("packagecode")

            for row in sheet.iter_rows(min_row=2, values_only=True):
                buyer_code = str(row[buyercode_idx]).strip() if row[buyercode_idx] else ""
                buyer_name = str(row[buyername_idx]).strip() if row[buyername_idx] else ""
                package_name = str(row[packagename_idx]).strip() if row[packagename_idx] else ""
                package_code = str(row[packagecode_idx]).strip() if row[packagecode_idx] else ""

                if not buyer_code or not buyer_name or not package_code:
                    print(f"Skipping row with missing values: {row}")
                    continue

                buyer = self.env["res.partner"].search([
                    ("code", "=", buyer_code),
                    ("tti_company_category", "=", "buyer")
                ], limit=1)

                if not buyer:
                    print(f"Buyer not found: {buyer_name} ({buyer_code}), skipping.")
                    continue

                package = self.env["product.template"].search([
                    ("default_code", "=", package_code),
                    ("name", "=", package_name)
                ], limit=1)

                if not package:
                    print(f"Package not found for {package_code}, skipping assignment.")
                    continue

                buyer.write({"tti_test_packages": [(4, package.id)]})
                print(f"Assigned Pakcage ({package_code}) to Buyer ({buyer_name} - {buyer_code})")
        except Exception as e:
            print(f"Error processing Excel file: {e}")

    @api.model
    def update_manuf_zones(self):
        # file_path = "custom_addons/2025/05_May/12_05_2025/1_tti_testing_v18/tti_testing/visio_tti_so_customize/update_data/zone_set.xlsx"
        file_path = "src/user/visio_tti_so_customize/update_data/zone_set.xlsx"

        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]

            required_fields = {"code", "city", "city_zone"}
            if not required_fields.issubset(set(headers)):
                _logger.error("Missing required columns in the sheet.")
                return

            code_idx = headers.index("code")
            city_idx = headers.index("city")
            city_zone_idx = headers.index("city_zone")

            # Pre-cache environment models
            Partner = self.env["res.partner"].sudo()
            City = self.env["tti.city"].sudo()
            CityZone = self.env["tti.city.zone"].sudo()

            for row in sheet.iter_rows(min_row=2, values_only=True):
                code = str(row[code_idx]).strip() if row[code_idx] else ""
                city = str(row[city_idx]).strip() if row[city_idx] else ""
                city_zone = str(row[city_zone_idx]).strip() if row[city_zone_idx] else ""

                if not code or not city:
                    _logger.warning(f"Skipping row with missing values: {row}")
                    continue

                manufacture = Partner.search([
                    ("code", "=", code),
                    ("company_id", "=", 2),
                    ("tti_company_category", "=", "manufacture")
                ], limit=1)

                if not manufacture:
                    _logger.warning(f"Manufacturer not found for code: {code} ({city}), skipping.")
                    continue

                city_id = City.search([
                    ("name", "=", city),
                    ("company_id", "=", 2)
                ], limit=1)

                if not city_id:
                    city_id = City.create({
                        "name": city,
                        "country_id": 177,
                        "company_id": 2
                    })

                city_zone_id = CityZone.search([
                    ("name", "=", city_zone),
                    ("tti_city_id", "=", city_id.id),
                    ("company_id", "=", 2)
                ], limit=1)

                if not city_zone_id:
                    city_zone_id = CityZone.create({
                        "name": city_zone,
                        "tti_city_id": city_id.id,
                        "company_id": 2
                    })

                manufacture.write({
                    "city": city,
                    "tti_city_id": city_id.id,
                    "tti_city_zone_id": city_zone_id.id,
                })

        except Exception as e:
            _logger.exception(f"Error processing Excel file: {e}")


class ProductTemplateUpdate(models.Model):
    _inherit = "product.template"

    @api.model
    def update_product_test_reports(self):
        file_path = "custom/addons/visio_tti_so_customize/update_data/tests data.xlsx"
        file_path = "src/user/visio_tti_so_customize/update_data/tests.xlsx"

        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]

            if "packagecode" not in headers or "testcode" not in headers or "qty" not in headers:
                _logger.error("Error: The sheet must have 'packagecode', 'testcode', and 'qty' columns.")
                return

            package_idx = headers.index("packagecode")
            test_idx = headers.index("testcode")
            qty_idx = headers.index("qty")

            packages = {}
            test_codes = set()
            package_codes = set()

            for row in sheet.iter_rows(min_row=2, values_only=True):
                package_code = str(row[package_idx]).strip() if row[package_idx] else ""
                test_code = str(row[test_idx]).strip() if row[test_idx] else ""
                if not package_code or not test_code:
                    # _logger.warning(f"Skipping row with missing values: {row}")
                    continue
                test_codes.add(test_code)
                package_codes.add(package_code)

            tests_map = {
                test.default_code: test.id
                for test in self.env["product.template"].search([
                    ("default_code", "in", list(test_codes)),
                    ("test_type", "=", "test_report")
                ])
            }

            for row in sheet.iter_rows(min_row=2, values_only=True):
                package_code = str(row[package_idx]).strip() if row[package_idx] else ""
                test_code = str(row[test_idx]).strip() if row[test_idx] else ""
                qty = int(row[qty_idx] or 0)

                if not package_code or not test_code or test_code not in tests_map.keys():
                    # _logger.warning(f"Skipping row with missing or invalid test code: {row}")
                    continue
                if qty <= 0:
                    qty = 1
                if package_code not in packages:
                    packages[package_code] = []
                packages[package_code].append((0, 0, {"test_report": tests_map[test_code], "qty": qty}))

            package_products = self.env["product.template"].search([
                ("default_code", "in", list(package_codes)),
                ("test_type", "=", "test_package")
            ])

            package_map = {package.default_code: package for package in package_products}
            counter = 0
            for package_code, tests in packages.items():
                package = package_map.get(package_code)
                if package:
                    counter+=1
                    package.write({"test_report_ids": tests})

            # _logger.info(f"Product Test Report update process completed. Total Packages : {len(packages)}, Add Packages : {len([*filter(None, packages.keys())])}")
            _logger.info(f"Product Test Report update process completed. Total Packages : {len(packages)}, Updated Packages : {counter}")
        except Exception as e:
            _logger.error(f"Error processing Excel file: {e}")

    @api.model
    def update_product_sale_price(self):
        file_path = "src/user/visio_tti_so_customize/update_data/product_price_update.xlsx"

        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active

            headers = [cell.value.strip() if cell.value else "" for cell in sheet[1]]
            if "code" not in headers or "price" not in headers:
                _logger.error("Error: The sheet must have 'code' and 'price' columns.")
                return

            code_idx = headers.index("code")
            price_idx = headers.index("price")

            # Fetch all relevant products in a single query
            product_records = self.env["product.template"].sudo().search([("default_code", "!=", False)])
            product_dict = {product.default_code: product for product in product_records}

            updates = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                code = row[code_idx]
                price = row[price_idx]

                if not code or price is None:
                    _logger.warning(f"Skipping row with missing values: {row}")
                    continue

                code = str(code).strip()
                try:
                    price = float(price)
                except ValueError:
                    _logger.warning(f"Invalid price format for product {code}: {price}, skipping.")
                    continue

                product = product_dict.get(code)
                if not product:
                    _logger.warning(f"Product not found: {code}, skipping.")
                    continue

                updates.append((product, price))

            # Perform batch updates
            for product, price in updates:
                product.write({"list_price": price})

            _logger.info("Product Sale Price update process completed.")

        except Exception as e:
            _logger.error(f"Error processing Excel file: {e}")

    @api.model
    def update_product_sale_price_usd(self):
        file_path = "src/user/visio_tti_so_customize/update_data/product_price_update.xlsx"

        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active

            headers = [cell.value.strip() if cell.value else "" for cell in sheet[1]]
            if "code" not in headers or "price" not in headers:
                _logger.error("Error: The sheet must have 'code' and 'price' columns.")
                return

            code_idx = headers.index("code")
            price_idx = headers.index("price")

            # Fetch all relevant products in a single query
            product_records = self.env["product.template"].sudo().search([("default_code", "!=", False)])
            product_dict = {product.default_code: product for product in product_records}

            updates = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                code = row[code_idx]
                price = row[price_idx]

                if not code or price is None:
                    _logger.warning(f"Skipping row with missing values: {row}")
                    continue

                code = str(code).strip()
                try:
                    price = float(price)
                except ValueError:
                    _logger.warning(f"Invalid price format for product {code}: {price}, skipping.")
                    continue

                product = product_dict.get(code)
                if not product:
                    _logger.warning(f"Product not found: {code}, skipping.")
                    continue

                updates.append((product, price))

            # Perform batch updates
            for product, price in updates:
                product.write({"list_price_usd": price})

            _logger.info("Product Sale Price update process completed.")

        except Exception as e:
            _logger.error(f"Error processing Excel file: {e}")
